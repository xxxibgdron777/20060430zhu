"""
飞书表格同步脚本
读取飞书电子表格数据，写入本地 Excel 文件
触发方式：Webhook 通知 / 手动调用
"""
import urllib.request
import urllib.error
import json
import os
import time
import shutil
import tempfile

# ==================== 飞书配置 ====================
try:
    from feishu_config import FEISHU_APP_ID, FEISHU_APP_SECRET, SPREADSHEET_TOKEN
except ImportError:
    # 部署前请复制 feishu_config.example.py 为 feishu_config.py 并填入真实值
    raise ImportError("未找到 feishu_config.py，请从 feishu_config.example.py 复制并填写")

# 目标文件路径（与 data_loader.py 共享同一文件）
TARGET_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "管理报表.xlsx")


def _get_token():
    """获取飞书 tenant_access_token"""
    data = json.dumps({"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET}).encode()
    req = urllib.request.Request(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        data=data,
        headers={"Content-Type": "application/json"}
    )
    resp = json.loads(urllib.request.urlopen(req).read())
    if resp.get("code") != 0:
        raise Exception(f"飞书认证失败: {resp.get('msg', resp)}")
    return resp["tenant_access_token"]


def _api_get(token, url):
    """飞书 GET 请求"""
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    resp = json.loads(urllib.request.urlopen(req).read())
    if resp.get("code") != 0:
        raise Exception(f"API 错误: {resp.get('code')} {resp.get('msg', '')}")
    return resp.get("data", {})


def sync():
    """
    同步飞书表格 → 本地 Excel
    返回: (success, message)
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False, "缺少 openpyxl 库，请安装: pip install openpyxl"

    t0 = time.time()
    try:
        token = _get_token()
        
        # 1. 获取表格元信息
        print("[feishu_sync] 获取表格结构...")
        meta = _api_get(token, f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{SPREADSHEET_TOKEN}/metainfo")
        title = meta.get("properties", {}).get("title", "管理报表")
        sheets = meta.get("sheets", [])
        print(f"[feishu_sync] 表格: {title}, {len(sheets)} 个 Sheet")

        # 2. 批量读取所有 Sheet 数据
        all_data = {}
        for sheet in sheets:
            sheet_id = sheet["sheetId"]
            sheet_title = sheet["title"]
            row_count = sheet["rowCount"]
            col_count = sheet["columnCount"]
            max_col_letter = get_column_letter(col_count)
            range_str = f"{sheet_id}!A1:{max_col_letter}{row_count}"

            print(f"[feishu_sync] 读取 [{sheet_title}] ({row_count}行)...")
            url = (f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                   f"{SPREADSHEET_TOKEN}/values/{urllib.request.quote(range_str)}")
            data = _api_get(token, url)
            values = data.get("valueRange", {}).get("values", [])
            
            if not values:
                print(f"[feishu_sync]  [{sheet_title}] 无数据，跳过")
                continue
            
            # 3. 写入 openpyxl
            print(f"[feishu_sync]  [{sheet_title}] 写入 {len(values)} 行...")
            all_data[sheet_title] = values

        # 4. 生成 Excel 文件（先写临时文件，避免中断时损坏原文件）
        import re
        def _eval_formula(s):
            """安全计算简单数学表达式，如 '865521.6/3'、'327490-3700'、'-630000/12'"""
            if not isinstance(s, str):
                return s
            s = s.strip()
            # 只允许数字、小数点、括号、四则运算符号
            if not re.match(r'^[\d\+\-\*\/\(\)\.\s]+$', s):
                return s
            try:
                result = eval(s, {"__builtins__": {}}, {})
                if isinstance(result, (int, float)):
                    return result
            except:
                pass
            return s

        def _clean_value(v):
            """转换飞书单元格值为 openpyxl 兼容格式"""
            if v is None:
                return None
            if isinstance(v, dict):
                return str(v.get('type', ''))
            if isinstance(v, (int, float, bool)):
                return v
            # 尝试计算公式字符串
            return _eval_formula(str(v))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认 Sheet

        for sheet_title, values in all_data.items():
            ws = wb.create_sheet(title=sheet_title[:31])  # Excel Sheet 名最多 31 字符
            for row_idx, row_data in enumerate(values, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=_clean_value(cell_value))

        # 5. 原子写入：先写临时文件，再替换
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(TARGET_FILE))
        os.close(tmp_fd)
        wb.save(tmp_path)
        shutil.move(tmp_path, TARGET_FILE)  # Linux 同分区是原子操作

        elapsed = round(time.time() - t0, 1)
        msg = f"同步成功：{len(all_data)} 个 Sheet，共 {sum(len(v) for v in all_data.values())} 行，耗时 {elapsed}秒"
        print(f"[feishu_sync] {msg}")
        return True, msg

    except Exception as e:
        elapsed = round(time.time() - t0, 1)
        msg = f"同步失败({elapsed}秒): {e}"
        print(f"[feishu_sync] {msg}")
        return False, msg


if __name__ == "__main__":
    ok, msg = sync()
    print(f"\n结果: {'✅' if ok else '❌'} {msg}")
